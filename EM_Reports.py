'''
    File name: EM_Reports.py
    Author: Jay Patel
    Date created: 1/12/2018
    Date last modified: 27/12/2018
    Python Version: 3.7.1
    Description: This program is intended to automate most of the process for creating the monthly Dashboard
    so as to reduce the time and effort that goes into downloading and processing various reports from
    EchoMobile. The program can currently get and process reports for any date range for any or all of the
    six following Groups: BMAMA, BMIMBA, MAMA, MIMBA, M2MAMA, M2MIMBA. As of now, only the following fields
    are included: phone, group, locationTextRaw, upload_date, and opted_out. A future version of the program
    may include options for the user to select a custom combination of fields.
'''

import base64
import csv
import datetime
import json
import os
import sys
import time
from collections import OrderedDict
from string import ascii_uppercase

import PySimpleGUI as sg
import pandas as pd
import requests
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
from openpyxl.formatting.rule import DataBarRule

# the following imports are required for pyinstaller:
# packaging, packaging.requirements, packaging.specifiers, packaging.version, six
import packaging
import packaging.requirements
import packaging.specifiers
import packaging.version
import six


# parameters required to generate a contact export
# type - an integer representing the report type
# target - the key of the group you want to export contacts
# std_field - these are comma separated contact properties you wish to include in the export, only include what you require
#           name - the name of the contact
#           phone - the phone number of the contact
#           internal_id - unique id for the contact that you've setup
#           group - the name of the group the contact belongs in
#           referrer - the name of the contact who referred the current contact
#           referrer_phone - the contact who referred the current contact
#           upload_date - the date the contact was created/uploaded
#           last_survey_complete_date - the date of the latest survey completion
#           geo - geocordinates of the contact's location
#           locationTextRaw - the contact's location name
#           labels - contact's labels
#           linked_entity - key to linked entity
#           opted_out - whether the contact is opted out
# field - comma separated custom field keys. Use this to include values of the contact's fields in the export
# sample params would be as shown below, please note that the keys used below are not valid, replace with your own

# MIMBA         ag1zfm0tc3dhbGktaHJkchMLEgtDbGllbnRHcm91cBjAmEQM
# BMAMA         ag1zfm0tc3dhbGktaHJkchMLEgtDbGllbnRHcm91cBi_3UQM
# BMIMBA        ag1zfm0tc3dhbGktaHJkchMLEgtDbGllbnRHcm91cBi88E4M
# M2MIMBA       ag1zfm0tc3dhbGktaHJkchMLEgtDbGllbnRHcm91cBii9VEM
# MAMA2         ag1zfm0tc3dhbGktaHJkchMLEgtDbGllbnRHcm91cBjEr2gM
# MAMA4         ag1zfm0tc3dhbGktaHJkchQLEgtDbGllbnRHcm91cBiMrKQRDA
# M2MAMA        ag1zfm0tc3dhbGktaHJkchQLEgtDbGllbnRHcm91cBizo4oSDA
# MIMBA2TEST    ag1zfm0tc3dhbGktaHJkchQLEgtDbGllbnRHcm91cBiT_pASDA
# MAMA          ag1zfm0tc3dhbGktaHJkchQLEgtDbGllbnRHcm91cBiGyIUTDA

# From Echo Mobile Support:
# Getting the Group Contacts report is done in two phases:
# In phase one, you initiate the export and receive the report key in the API response
# In phase two, you poll the system via an API until the export has been concluded

class EMReports:

    def __init__(self):

        self.eid = '1015527'
        self.epassword = '81ldtknyqgy6bl0'
        self.account_id = b'5338310569623552'
        self.AUTH_METHOD = 'ENTERPRISE'  # or ACCOUNT, see https://www.echomobile.org/public/api/authentication
        self.CONTACT_REPORT_TYPE = 12

        # Suppress SettingWithCopyWarning in process_report_into_excel function
        pd.options.mode.chained_assignment = None  # default='warn'

    def get_headers(self):
        emr = EMReports()

        # Option 1 - enterprise authentication - PREFERRED
        # see https://www.echomobile.org/app/organization/settings under advanced/developer options
        if emr.AUTH_METHOD == 'ENTERPRISE':
            password = emr.epassword
            uname = emr.eid

        # encode the account id to b64 and get the string of that value because that's what the EM servers expect
        account_id_bytes = base64.b64encode(emr.account_id)
        account_id_str = account_id_bytes.decode('utf-8')

        headers = {
            'authorization': emr.basic_auth_header(uname, password),
            'account-id': account_id_str,
        }

        return headers

    def basic_auth_header(self, user, password):
        # encode the password to b64 and get the string of that value because that's what the EM servers expect
        user_password = '{}:{}'.format(user, password).encode('utf-8')
        base64_user_password_bytes = base64.b64encode(user_password)
        base64_user_password = base64_user_password_bytes.decode('utf-8')

        return 'Basic {}'.format(base64_user_password)

    def start_generate_report(self, params, output_filename='report.csv'):
        emr = EMReports()

        _type = params['type']
        print('Generating Contacts Report')
        r = requests.post(
            'https://www.echomobile.org/api/cms/report/generate',
            data=params,
            headers=emr.get_headers()
        )

        response = json.loads(r.text)
        return emr.poll_report(response.get('rkey'), output_filename)

    def poll_report(self, report_key, output_filename):
        emr = EMReports()

        # for very large exports you should increase the max poll attempts so that we don't time out prematurely
        MAX_POLL_ATTEMPTS = 200
        DELAY_BETWEEN_POLL = 5  # seconds
        if not report_key:
            print('Error: Missing report key!')
            return

        attempts = 0
        while True:
            print('Attempting to download report to {}, attempt {}'.format(output_filename, attempts + 1))
            r = requests.get(
                'https://www.echomobile.org/api/cms/report/serve?rkey={}'.format(report_key),
                headers=emr.get_headers()
            )

            attempts += 1

            if r.text == 'Unauthorized' and attempts < MAX_POLL_ATTEMPTS:
                # the api currently returns unauthorized if the generation process in not complete, check again later
                time.sleep(DELAY_BETWEEN_POLL)
            elif attempts > MAX_POLL_ATTEMPTS:
                print('Error: Timed out trying to fetch report, try setting a higher MAX_POLL_ATTEMPTS!')
                break
            else:
                break  # we have the report

        # We now have contents of the report in csv format, to do with whatever you please
        return r.text

    # get reports from EM server, save in .CSV format to folder "CSVDict"
    def get_report(self, group_dict, user_selection):

        emr = EMReports()
        get_these_reports = {}
        now_str = datetime.datetime.now().strftime("%Y%m%d_%H-%M-%S")

        for key, val in group_dict.items():
            if key in user_selection[0]:
                get_these_reports[key] = val

        CSV_Dir = 'C:\\JHSL\\EM_reports\\CSV\\{}'.format(now_str)
        for report_name, group_key in get_these_reports.items():
            out_file = '{}-{}.csv'.format(report_name, now_str)
            params = {
                'type': emr.CONTACT_REPORT_TYPE,
                'target': group_key,  # replace with your group key
                'std_field': 'phone, group, locationTextRaw, upload_date, opted_out'
            }

            print('Getting {} report...'.format(report_name))
            output_csv = emr.start_generate_report(params, output_filename=out_file)

            if not os.path.exists(CSV_Dir):
                os.makedirs(CSV_Dir)

            file_path = os.path.join(CSV_Dir, out_file)
            print('Downloaded report, saving to file {}'.format(file_path))
            print()

            with open(file_path, 'w') as file:
                file.write(output_csv)

        return CSV_Dir, now_str

    # read each csv file, sort by date, discard unwanted dates, save each csv in separate worksheet in excel file,
    # combine all data and save into another worksheet
    def process_report_into_excel(self, user_input, CSV_Dir):

        start_date = pd.to_datetime(user_input[1])
        end_date = pd.to_datetime(user_input[2])
        reports = {}
        out_xl_filename = 'temp.xlsx'

        for file in os.listdir(CSV_Dir):
            if file.endswith('.csv'):
                group = file.split('-')[0]
                csv_filepath = os.path.join(CSV_Dir, file)
                df = pd.read_csv(csv_filepath)
                df['upload_date'] = pd.to_datetime(df['upload_date'])
                df2 = df.sort_values('upload_date')

                # pandas magic right here
                # https://erikrood.com/Python_References/select_pandas_dataframe_rows_between_two_dates_final.html
                mask = (df2['upload_date'] > start_date) & (df2['upload_date'] < end_date)
                df3 = df2.loc[mask]

                reports[group] = df3

        xl_filepath = os.path.join(CSV_Dir, out_xl_filename)
        xl_writer = pd.ExcelWriter(xl_filepath)
        all_data_list = list(reports.values())  # get all values from reports dict in a list

        for report_name, report_data in reports.items():
            # CSV from EM API doesn't include group name (check this with the support team)
            # The line below will fill in the group name where necessary
            # However, it returns SettingWithCopyWarning: "A value is trying to be set on a copy of a slice from a DataFrame
            # Try using .loc[row_indexer,col_indexer] = value instead"
            # Warning has been suppressed
            report_data['group'] = report_name
            report_data.to_excel(xl_writer, report_name)

        all_reports_data = pd.concat(all_data_list)  # merge all data into one dataframe
        date_sorted_all_reports = all_reports_data.sort_values('upload_date')  # sort by date
        date_sorted_all_reports.to_excel(xl_writer, 'COMBINED_REPORTS')
        xl_writer.save()

        return xl_filepath

    # find column titled 'locationTextRaw' and insert a column directly after it in which the fuzzy matches
    # will be inserted. Also find and return the 'upload date' column for use in analysis
    def column_actions(self, wb):
        alphabet_upper = list(ascii_uppercase)

        for ws in wb.worksheets:
            ws.delete_cols(1)
            ws.insert_cols(4)

            last_in_row = '{}{}'.format(ascii_uppercase[len(ws[1])], '1')

            headers = ws['A1':last_in_row]  # look in the range
            for col in headers:
                for cell in col:
                    if cell.value is not None:
                        if 'locationTextRaw' in cell.value:
                            fuzzy_title_col = str(alphabet_upper[alphabet_upper.index(cell.column) + 1])
                        elif 'upload_date' in cell.value:
                            upload_date_col = str(alphabet_upper[alphabet_upper.index(cell.column)])

            new_fuzzy_title_index = fuzzy_title_col + '1'
            ws[new_fuzzy_title_index] = 'Location Match'

        return wb, fuzzy_title_col, upload_date_col

    # fuzzy match raw location against (modified) DHIS2 facility list
    def fuzzy_match(self, file, CSV_Dir):
        emr = EMReports()
        wb = load_workbook(file)
        wb, fuzzy_title_col, upload_date_col = emr.column_actions(wb)
        exclude_list = ['none', 'home', 'house', 'nyumbani', 'Nyumbani', 'Hospitali', 'bmama', 'bmimba', 'm2mama', 'm2mimba', 'mama', 'mimba', 'Nairobi', 'nairobi']

        facility_file = 'C:\\JHSL\\EM_reports\\parameters\\main_facility_list.csv'
        with open(facility_file, mode='r') as infile:
            reader = csv.reader(infile)
            facility_dict = dict(reader)

        best_match_dict = {}
        temp_match_dict = {}

        # TODO: optimize concat of sheets to 'combined sheets' after fuzzy match if it adversely affects performance

        raw_location_exclude_list = ['yes', 'ndio', 'hospital', 'hospitali', 'hosipitali', 'centre', 'center', 'services', 'clinic']
        # 'sub', 'district']

        # iterate through worksheets, match each raw_location against entire facility list, write the best match if match ratio >= 10
        for sheet in wb.worksheets:
            print('Fuzzy Matching {}'.format(sheet))
            for each_row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):  # iterate through col C
                raw_loc = each_row[0].value

                try:
                    raw_location_lower = raw_loc.lower()
                except TypeError:
                    raw_location_lower = raw_loc

                for word in raw_location_exclude_list:
                    if word in raw_location_lower:
                        raw_location_lower = raw_location_lower.replace(word, '')

                # eliminate strings which are too long and too short
                if len(raw_location_lower) > 3 and len(raw_location_lower) <= 45:
                    if raw_location_lower is not None and raw_location_lower is not '--':

                        # the statement below is to counteract weird matching results
                        # and is far from being optimized
                        # TODO: the above needs to be done
                        if raw_location_lower == 'chwele':
                            raw_location_lower = 'chwele sub district hospital'
                        elif raw_location_lower == 'chwele hospital':
                            raw_location_lower = 'chwele sub district hospital'
                        elif raw_location_lower == 'lugulu' or raw_location_lower == 'lukulu' or raw_location_lower == 'lugulu hospital':
                            raw_location_lower = 'lugulu friends mission hospital'
                        elif raw_location_lower == 'thika' or raw_location_lower == 'thika level 5 hospital' or raw_location_lower == 'level 5 thika':
                            raw_location_lower = 'thika level 5'
                        elif raw_location_lower == 'bungoma hospital':
                            raw_location_lower = 'bungoma county referral'
                        elif raw_location_lower == 'kiandutu':
                            raw_location_lower = 'kiandutu health centre'
                        elif raw_location_lower == 'kiambu' or raw_location_lower == 'kiambu hospital':
                            raw_location_lower = 'kiandutu health centre'
                        elif raw_location_lower == 'tigoni':
                            raw_location_lower = 'tigoni district hospital'
                        elif raw_location_lower == 'bungoma hospital':
                            raw_location_lower = 'bungoma county referral hospital'
                        elif raw_location_lower == 'kihara':
                            raw_location_lower = 'kihara sub-county hospital'
                        elif raw_location_lower == 'kihara':
                            raw_location_lower = 'kihara sub-county hospital'
                        elif raw_location_lower == 'kanganga':
                            raw_location_lower = 'kanganga dispensary'
                        elif raw_location_lower == 'wagige':
                            raw_location_lower = 'wangige sub county hospital'
                        elif raw_location_lower == 'wangige':
                            raw_location_lower = 'wangige sub county hospital'
                        elif 'mama' in raw_location_lower and 'lucy' not in raw_location_lower and 'margaret' not in raw_location_lower:
                            raw_location_lower = ''
                        elif raw_location_lower == 'mother':
                            raw_location_lower = ''
                        elif raw_location_lower in exclude_list:
                            raw_location_lower = ''

                        for each_facility in facility_dict.keys():  # iterate through entire facility file
                            fuzz_ratio = fuzz.partial_ratio(raw_location_lower,
                                                            each_facility)  # get partial match ratio
                            temp_match_dict[each_facility] = fuzz_ratio  # facility: ratio

                        # get highest value in temp_match_dict and return key
                        # https://stackoverflow.com/questions/268272/getting-key-with-maximum-value-in-dictionary
                        best_match = max(temp_match_dict, key=temp_match_dict.get)

                        best_match_dict[best_match] = fuzz_ratio  # add best match to dict

                    if best_match_dict[best_match] >= 10:  # only add matches with a ratio greater than 10
                        cell_index = fuzzy_title_col + str(each_row[0].row)
                        sheet[cell_index] = facility_dict[best_match]  # + " " + str(fuzz_ratio)  # write to cell

                    # print('{}: {}: {}'.format(raw_location_lower, best_match, fuzz_ratio))  # leave for debugging

        # wb.save(out_filepath)
        return wb

    # write to worksheet
    def write_analysis_to_worksheet(self, ws, months_dict, date_list, boolean_dict):
        date_dict = OrderedDict(date_list)

        # "enr" = enrolment
        daily_enr_col_num = ws.max_column + 2
        daily_max_rows = len(date_dict) + 2
        sum_daily = sum(date_dict.values())
        max_daily = max(date_dict.values())

        monthly_enr_col_num = daily_enr_col_num + 3
        monthly_max_rows = len(months_dict) + 1

        boolean_row = monthly_enr_col_num + 3

        # write DAILY  errolment numbers to worksheet in the appropriate columns
        for daily_enr_row in ws.iter_rows(min_row=1, max_row=daily_max_rows, min_col=daily_enr_col_num,
                                          max_col=daily_enr_col_num + 1):
            daily_col1_index = '{}{}'.format(daily_enr_row[0].column,
                                             daily_enr_row[0].row)  # get cell col and row as index while iterating
            daily_col2_index = '{}{}'.format(daily_enr_row[1].column, daily_enr_row[1].row)

            if daily_enr_row[0].row == daily_max_rows:  # write total if last row
                ws[daily_col1_index] = 'TOTAL'
                ws[daily_col2_index] = sum_daily
            else:
                if daily_enr_row[0].row == 1:  # write header if first row
                    ws[daily_col1_index] = 'DATE'
                    ws[daily_col2_index] = '# OF ENROLMENTS'
                    daily_first_index = daily_col2_index
                else:  # write values
                    daily_popped_val = date_dict.popitem(False)
                    ws[daily_col1_index] = daily_popped_val[0]
                    ws[daily_col2_index] = daily_popped_val[1]
                    daily_last_index = daily_col2_index

        # order according to month if multiple months
        months_dict_ordered = OrderedDict(sorted(months_dict.items()))

        # write MONTHLY enrolment numbers to worksheet in the appropriate columns
        for monthly_enr_row in ws.iter_rows(min_row=1, max_row=monthly_max_rows, min_col=monthly_enr_col_num,
                                            max_col=monthly_enr_col_num + 1):
            monthly_col1_index = '{}{}'.format(monthly_enr_row[0].column,
                                               monthly_enr_row[0].row)  # get cell col and row as index while iterating
            monthly_col2_index = '{}{}'.format(monthly_enr_row[1].column, monthly_enr_row[1].row)

            if monthly_enr_row[0].row == 1:  # write header if first row
                ws[monthly_col1_index] = 'MONTH'
                ws[monthly_col2_index] = '# OF ENROLMENTS'
            else:  # write values
                monthly_popped_val = months_dict_ordered.popitem(False)
                ws[monthly_col1_index] = monthly_popped_val[0]
                ws[monthly_col2_index] = monthly_popped_val[1]

        # conditional formatting bars
        rule = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=max_daily, color="FF638EC6",
                           showValue="None", minLength=None, maxLength=None)
        cell_range = '{}:{}'.format(daily_first_index, daily_last_index)
        ws.conditional_formatting.add(cell_range, rule)

        for boolean_count_row in ws.iter_rows(min_row=1, max_row=3, min_col=boolean_row,
                                              max_col=boolean_row + 1):
            boolean_col1_index = '{}{}'.format(boolean_count_row[0].column,
                                               boolean_count_row[
                                                   0].row)  # get cell col and row as index while iterating
            boolean_col2_index = '{}{}'.format(boolean_count_row[1].column, boolean_count_row[1].row)

            if boolean_count_row[0].row == 1:  # write header if first row
                ws[boolean_col1_index] = 'OPTED OUT'
                ws[boolean_col2_index] = 'NUMBER'
            else:  # write values
                boolean_popped_val = boolean_dict.popitem()
                ws[boolean_col1_index] = boolean_popped_val[0]
                ws[boolean_col2_index] = boolean_popped_val[1]

        """chart1 = BarChart()
        chart1.type = "col"
        chart1.shape = 4
    
        data = Reference(ws, min_col=8, max_col=9, min_row=2, max_row=30)
    
        chart1.add_data(data)
        ws.add_chart(chart1, "R1")"""

    # do some basic analysis and return workbook
    def analysis(self, wb, start_date):
        print()
        print('Analysing...')
        emr = EMReports()

        # TODO: count number and percentage of opts out

        # iterate through upload_date column and count total number of occurrences of enrolments on each date, month, and per clock hour
        for ws in wb.worksheets:
            months_dict = {}
            date_dict = {}
            # hours_dict = dict.fromkeys(list(range(25)), 0)
            date_list = []
            header_list = []
            boolean_dict = {True: 0, False: 0}

            for cell in ws[1]:
                header_list.append(cell.value)

            date_col_number = header_list.index('upload_date') + 1
            opted_out_col_number = header_list.index('opted_out') + 1

            for row1 in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=date_col_number, max_col=date_col_number):
                cell_value1 = row1[0].value  # get value of cell

                if isinstance(cell_value1, datetime.date):  # process only if cell value is a datetime object

                    # count number of enrolments per day
                    each_date = cell_value1.date()
                    if each_date not in date_dict:
                        date_dict[each_date] = 1
                    else:
                        date_dict[each_date] = date_dict[each_date] + 1

                    # count number of enrolments per each_month
                    each_month = cell_value1.date().month
                    if each_month not in months_dict:
                        months_dict[each_month] = 1
                    else:
                        months_dict[each_month] = months_dict[each_month] + 1

                    """# count number of enrolments per clock hour
                    each_hour = cell_value1.time().hour
                    hours_dict[each_hour] = hours_dict[each_hour] + 1"""

            for row2 in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=opted_out_col_number,
                                     max_col=opted_out_col_number):
                cell_value2 = row2[0].value  # get value of cell
                if cell_value2 is True:
                    boolean_dict[True] += 1
                elif cell_value2 is False:
                    boolean_dict[False] += 1

            # convert dict to list of tuples so as to allow for sorting by date
            # Future: sort the list in place as per the link below
            # https://www.saltycrane.com/blog/2007/09/how-to-sort-python-dictionary-by-keys/
            for key, value in date_dict.items():
                date_list.append((key, value))

            first_date = start_date - datetime.timedelta(1)
            break_flag = False

            # if any days are missing from the spreadsheet, the missing days will be added in along with an enrolment number of zero
            for each_day in sorted(date_list):
                difference = each_day[0] - first_date
                # if difference between current day and previous day is greater than 1, i.e. if any days are missing
                if difference > datetime.timedelta(1):
                    date_list.append(each_day[0] + datetime.timedelta(1))
                    for i in range(difference.days):
                        if i > 0:
                            date_to_append = (first_date + datetime.timedelta(i), 0)
                            date_list.append(date_to_append)
                            break_flag = True

                if break_flag is True:
                    first_date += datetime.timedelta(i + 1)
                    continue
                else:
                    first_date += datetime.timedelta(1)

            for i in date_list:
                if not isinstance(i, tuple):
                    date_list.remove(i)

            date_list.sort()

            emr.write_analysis_to_worksheet(ws, months_dict, date_list, boolean_dict)

        return wb


class FileActions:

    def __init__(self):
        pass

    # read groups ino from file
    def read_params(self, current_dir, filename):

        filepath = os.path.join(current_dir, filename)
        groups_dict = {}

        with open(filepath) as f:
            for line in f:
                if not line.startswith('#'):
                    (key, val) = line.split()
                    groups_dict[key] = val

        groups_list = list(groups_dict.keys())

        return groups_dict, groups_list


class GUI:

    # launch GUI
    def start_gui(self, group_list):

        gui = GUI()
        checkbox_list = []
        report_list_upper = [x.upper() for x in group_list]
        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

        for each in report_list_upper:
            checkbox_list.append(sg.Checkbox(each))

        # TODO: build in date validation and ensure at least one box is checked

        layout = [
            [sg.Text('JHSL Reports Assistant for Echo Mobile', size=(45, 1), font=("Helvetica", 15), text_color='green',
                     justification='Center')],
            [sg.Text('')],
            [sg.Text('Please Select The Reports You Want', size=(35, 1), font=("Helvetica", 15), text_color='blue4')],
            checkbox_list,
            [sg.Text('==============================================================================================')],
            [sg.Text('Please Enter the Start and End Dates', size=(35, 1), font=("Helvetica", 15), text_color='blue4')],
            [sg.In('', size=(20, 1), key='start_date')],
            [sg.CalendarButton('Start Date', target='start_date', key='start_date')],
            [sg.In('', size=(20, 1), key='end_date')],
            [sg.CalendarButton('End Date', target='end_date', key='end_date')],
            [sg.Text('==============================================================================================')],
            [sg.Text('Please Select the Output Folder', size=(35, 1), font=("Helvetica", 15), text_color='blue4')],
            [sg.InputText(desktop), sg.FolderBrowse()],
            [sg.Text('==============================================================================================')],
            [sg.Submit(), sg.Cancel()]]

        window = sg.Window('REPORT PROCESSOR').Layout(layout)
        event, values = window.Read()
        if event is 'Cancel':
            sys.exit(0)

        window.Close()

        # convert string from gui to datetime object
        datetime_start = datetime.datetime.strptime(values['start_date'], '%Y-%m-%d %H:%M:%S')
        datetime_end = datetime.datetime.strptime(values['end_date'], '%Y-%m-%d %H:%M:%S')

        # change to pandas timestamp date object, increment end date by 1
        pd_start_date = pd.Timestamp(datetime_start).date()
        pd_end_date = pd.Timestamp(datetime_end).date() + pd.Timedelta(days=1)

        checkbox_selection = gui.get_input(values, report_list_upper)
        user_input = [checkbox_selection, pd_start_date, pd_end_date]
        output_folder = values[6]

        return user_input, output_folder

    # extract and return user selected groups from gui input
    def get_input(self, values, group_list):
        key_list = []
        select_list = []

        # get keys for which the value is "True" and...
        for key, val in values.items():
            if val is True:
                key_list.append(key)

        # ...match with corresponding index in group_list
        for each in key_list:
            select_list.append(group_list[each])

        return select_list


class Engage:
    def __init__(self):
        pass

    def engage(self):
        params_dir = 'C:\\JHSL\\EM_reports\\parameters\\'
        filename = 'group_details-DO-NOT-DELETE-ME.txt'

        emr = EMReports()
        gui = GUI()
        fa = FileActions()

        group_dict, group_list = fa.read_params(params_dir, filename)

        # GUI Input
        user_input, output_folder = gui.start_gui(group_list)

        # start timer
        start_time = time.time()

        # Output CSVs
        csv_dir, output_filename = emr.get_report(group_dict, user_input)

        # Read CSVs and process
        file = emr.process_report_into_excel(user_input, csv_dir)

        # Fuzzy match raw locations agains DHIS2 facility list
        wb = emr.fuzzy_match(file, csv_dir)

        # Engage Analysis
        wb2 = emr.analysis(wb, user_input[1])

        print()
        ouput_filepath = '{}\\{}.xlsx'.format(output_folder, output_filename)
        wb2.save(ouput_filepath)
        print('Complete!')
        print('Output File saved to: {}'.format(ouput_filepath))

        # end timer
        print()
        print("--- {} seconds ---".format(time.time() - start_time))
        print('This window will self-destruct in 10 seconds!')
        time.sleep(10)
        sys.exit(0)


if __name__ == '__main__':
    do = Engage()
    do.engage()
